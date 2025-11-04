# apps/dashboard/utils.py
from datetime import datetime, timedelta
from decimal import Decimal

from django.utils.timezone import make_aware, now
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.worksheet.worksheet import Worksheet

from apps.bookings.models import Booking
from apps.common.choices import BookingStatus, PaymentMarker
from apps.bookings.admin import _calc_total_sum, _calc_commission, _calc_partner_net


def parse_period(request):
    """
    Возвращает (date_from, date_to) как aware datetime.
    Если не передан ?from=YYYY-MM-DD&to=YYYY-MM-DD,
    то период = последние 30 дней.
    """
    now_aware = now()

    df_default = now_aware - timedelta(days=30)
    dt_default = now_aware

    def parse_one(key, default_start, end=False):
        """
        key='from' или 'to'.
        Ждём формат '2025-10-01'.
        Возвращаем aware datetime.
        Если end=True -> конец дня 23:59:59.
        """
        val = request.GET.get(key)
        if not val:
            dt = default_start
        else:
            naive = datetime.strptime(val, "%Y-%m-%d")
            naive = naive.replace(
                hour=23 if end else 0,
                minute=59 if end else 0,
                second=59 if end else 0,
            )
            dt = make_aware(naive)  # tz возьмётся из TIME_ZONE
        return dt

    date_from = parse_one("from", df_default, end=False)
    date_to = parse_one("to", dt_default, end=True)
    return date_from, date_to


def base_queryset(date_from, date_to, partner_id=None, region_id=None):
    """
    Базовый queryset для отчёта.
    Включаем брони, которые реально приносят деньги:
    confirmed / issued / completed.
    rejected / expired / canceled — не считаем в выручку.
    """
    qs = (
        Booking.objects.filter(
            date_from__lte=date_to,
            date_to__gte=date_from,
            status__in=[
                BookingStatus.CONFIRMED,
                BookingStatus.ISSUED,
                BookingStatus.COMPLETED,
            ],
        )
        .select_related("car", "car__region", "partner")
    )

    if partner_id:
        qs = qs.filter(partner_id=partner_id)

    if region_id:
        qs = qs.filter(car__region_id=region_id)

    return qs


def calc_stats(qs):
    """
    Считаем KPI + строим данные для двух рейтингов:
    - Топ авто по обороту
    - Топ партнёров по обороту

    Возвращаем dict с:
      total_bookings, total_turnover, total_paid,
      total_commission, total_partner_net,
      top_cars_raw -> [(car_name, Decimal(sum)), ...],
      top_partners_raw -> [(partner_name, Decimal(sum)), ...],
      chart_labels / chart_values            -> машины
      partners_chart_labels / partners_chart_values -> партнёры
    """
    total_bookings = qs.count()

    total_turnover = Decimal("0")
    total_paid = Decimal("0")
    total_commission = Decimal("0")
    total_partner_net = Decimal("0")

    by_car = {}
    by_partner = {}

    for b in qs:
        sum_total = Decimal(str(_calc_total_sum(b)))
        sum_comm = Decimal(str(_calc_commission(b)))
        sum_net = Decimal(str(_calc_partner_net(b)))

        total_turnover += sum_total
        total_commission += sum_comm
        total_partner_net += sum_net

        if b.payment_marker == PaymentMarker.PAID:
            total_paid += sum_total

        car_name = b.car.title if b.car else "—"
        by_car[car_name] = by_car.get(car_name, Decimal("0")) + sum_total

        partner_name = b.partner.name if b.partner else "—"
        by_partner[partner_name] = by_partner.get(partner_name, Decimal("0")) + sum_total

    top_cars_items = sorted(by_car.items(), key=lambda x: x[1], reverse=True)[:5]
    top_partners_items = sorted(by_partner.items(), key=lambda x: x[1], reverse=True)[:5]

    cars_labels = [name for (name, _sum) in top_cars_items]
    cars_values = [float(_sum) for (_, _sum) in top_cars_items]

    partners_labels = [name for (name, _sum) in top_partners_items]
    partners_values = [float(_sum) for (_, _sum) in top_partners_items]

    return {
        "total_bookings": total_bookings,
        "total_turnover": total_turnover,
        "total_paid": total_paid,
        "total_commission": total_commission,
        "total_partner_net": total_partner_net,
        "top_cars_raw": top_cars_items,
        "top_partners_raw": top_partners_items,
        "chart_labels": cars_labels,
        "chart_values": cars_values,
        "partners_chart_labels": partners_labels,
        "partners_chart_values": partners_values,
    }


# ---------- XLSX helpers ----------

_THIN = Side(border_style="thin", color="D9D9D9")
BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
HEAD_FILL = PatternFill("solid", fgColor="F2F4F7")
HEAD_FONT = Font(bold=True)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")


def _auto_width(ws: Worksheet, cols: int):
    for col_idx in range(1, cols + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[col_letter]:
            try:
                v = str(cell.value) if cell.value is not None else ""
            except Exception:
                v = ""
            max_len = max(max_len, len(v))
        ws.column_dimensions[col_letter].width = min(max(12, max_len + 2), 40)


def build_excel(qs, date_from, date_to):
    """
    Готовим .xlsx отчёт (1 лист):
    - Заголовок
    - Период
    - Таблица детальности
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    # 1) Шапка
    title = "Отчёт по бронированиям"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=11)
    c = ws.cell(row=1, column=1, value=title)
    c.font = Font(size=14, bold=True)
    c.alignment = CENTER

    # 2) Период
    period_text = f"Период: {date_from.strftime('%Y-%m-%d')} \u2192 {date_to.strftime('%Y-%m-%d')}"
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=11)
    c2 = ws.cell(row=2, column=1, value=period_text)
    c2.alignment = CENTER
    c2.font = Font(color="666666")

    ws.append([])  # пустая строка после шапки (это станет строкой 3)

    # 3) Заголовки таблицы
    headers = [
        "Booking ID",
        "Дата с",
        "Дата по",
        "Машина",
        "Регион",
        "Партнёр",
        "Статус",
        "Сумма аренды (UZS)",
        "Комиссия агрегатора (UZS)",
        "Чистыми партнёру (UZS)",
        "Оплата",
    ]
    ws.append(headers)

    # применяем стиль к заголовкам
    head_row = 4
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=head_row, column=col_idx)
        cell.font = HEAD_FONT
        cell.fill = HEAD_FILL
        cell.alignment = CENTER
        cell.border = BORDER

    # 4) Данные
    money_format = "#,##0"  # openpyxl применит формат локально в Excel
    for b in qs:
        region_name = b.car.region.name if (b.car and b.car.region) else ""
        total_sum = _calc_total_sum(b)
        commission = _calc_commission(b)
        partner_net = _calc_partner_net(b)

        row = [
            b.id,
            b.date_from.strftime("%Y-%m-%d %H:%M"),
            b.date_to.strftime("%Y-%m-%d %H:%M"),
            b.car.title if b.car else "",
            region_name,
            b.partner.name if b.partner else "",
            b.get_status_display(),
            float(total_sum),   # как число
            float(commission),  # как число
            float(partner_net), # как число
            "Оплачено" if b.payment_marker == PaymentMarker.PAID else "Не оплачено",
        ]
        ws.append(row)

        # стили для только что добавленной строки
        r = ws.max_row
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=col_idx)
            cell.alignment = LEFT if col_idx not in (1,) else CENTER
            cell.border = BORDER

        # числовой формат для сумм
        ws.cell(row=r, column=8).number_format = money_format
        ws.cell(row=r, column=9).number_format = money_format
        ws.cell(row=r, column=10).number_format = money_format

    # автоширина
    _auto_width(ws, len(headers))

    return wb
